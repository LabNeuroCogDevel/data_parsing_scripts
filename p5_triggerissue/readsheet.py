#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime, pandas, sys, os.path
from openpyxl import load_workbook

fname='/Volumes/B/bea_res/Schizophrenia/Schizophrenia_scan_log.xlsx'
wsname='All Scans'

#if len(sys.argv) !=3:
#  print("USAGE: %s file sheet"%sys.argv[0])
#  sys.exit(1)
#
#fname=sys.argv[1]
#wsname=sys.argv[2]

if not os.path.isfile(fname): 
  print("'%s' is not a file!"%fname)
  sys.exit(1)


wb=load_workbook(filename=fname)

if wsname not in wb.sheetnames:
  print("'%s' not in '%s'; options:"%(wsname,fname))
  for n in wb.sheetnames: print("\t%s"%n)
  sys.exit(1)

sheet=wb[wsname]
# toprow = [ x for x in sheet['A1:ZZ1'] ][0] # pre openpyxl 2.4
toprow=sheet['A1:ZZ1'][0] 
# make sheet int a datatable
colnames= [x.value for x in toprow if x.value ] 

sheet=wb[wsname]
df = pandas.DataFrame( sheet.values )
# remove header row, and only grab columns with names
d = df.ix[1:,0:(len(colnames)-1)]
d.columns = colnames

#d.to_csv(sys.stdout,sep="\t",index=False)
triggerissueidx = [ True if re.search("trigger",x or " " ) else False for x in d['Notes'] ]
# all 12 with "trigger problems" had eyetracking
d['Eyetracking?'][ triggerissueidx ].values

